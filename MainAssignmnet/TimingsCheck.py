import openpyxl


def workbook(filename, sheetname):
    wb = openpyxl.load_workbook(filename)
    sh1 = wb[sheetname]
    row = sh1.max_row
    col = sh1.max_column
    return row, col, sh1, wb


row, col, sh1, wb=workbook("forTimings.xlsx","Sheet1")
next=sh1.cell(2,8).value
showLength=sh1.cell(2,3).value
sum=next+showLength
sh1.cell(2,12,value=sum)
wb.save("forTimings.xlsx")
no_of_shows=sh1.cell(2,7).value
print(no_of_shows)
print(col)
for i in range(1,no_of_shows+1):
    for column in range(13,col+1):
        prev=sh1.cell(2,column-1).value
        len=sh1.cell(2,3).value
        sum=prev+len
        sh1.cell(2, column, value=sum)
    no_of_shows = sh1.cell(2, 7).value
wb.save("forTimings.xlsx")
timingprev=sh1.cell(2, 8).value
timingnext = sh1.cell(2, 12).value
print(f"1.{timingprev}-{timingnext}")
i = 2
for ticketsbook in range(13, col + 1):
    timingprev=sh1.cell(2, ticketsbook-1).value
    timingnext = sh1.cell(2, ticketsbook).value
    print(f"{i}.{timingprev}-{timingnext}")
    i=i+1







# result_cell = 'L{}'.format(2)
# add_function = '=H{}+C{}'.format(2)
# sh1[result_cell] = add_function
# sh1.cell(2,12,value=sh1.cell(2,8)"to" (sh1.cell(2,8)+sh1.cell(2,3)))
#successfull
# b_col, c_col ,f_col = ['H', 'C', 'G' ]
# for row in range(2, 3):
#     result_cell = 'L{}'.format(row)
#
#     b_value = sh1[b_col + str(row)].value
#     c_value = sh1[c_col +str(row)].value
#     d_value=float(sh1.cell(2,8).value)
#
#
#     sh1[result_cell] =(b_value + c_value)
# for row in range(2,3):
#     result_cell='M{}'.format(row)
#     next=sh1.cell(2,12).value
#     showLength=sh1.cell(2,3).value
#     sum=next+showLength
#     sh1[result_cell]=sum
#
#
# wb.save("forTimings.xlsx")
# print(f"{sh1.cell(2,8).value}-{sh1.cell(2,12).value}")
# print(f"{sh1.cell(2,12).value}-{sh1.cell(2,13).value}")

