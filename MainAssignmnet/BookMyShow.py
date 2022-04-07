import sys
import openpyxl


def workbook(filename, sheetname):
    wb = openpyxl.load_workbook(filename)
    sh1 = wb[sheetname]
    row = sh1.max_row
    col = sh1.max_column
    return row, col, sh1, wb


class AdminOfBookMyShow():

    def __init__(self, adminName, adminPassword):
        self.adminName = adminName
        self.adminPassword = adminPassword


    def adminOperations(self):
        while (True):
            print("******WelcomeAdmin*******\n1.Add New Movie Info\n2.Edit Movie\n3.Delete Movie\n4.Logout")
            choice=int(input("Enter the choice you want to perform"))
            row, col, sh1, wb = workbook("MoviesInfo.xlsx", "Sheet1")
            if(choice==4):
                break
            if (choice == 1):
                print("enter the details of the movie: ")
                row1 = 1
                column1 = 1
                for c in range(1, col + 1):
                    m = (sh1.cell(row1, column1).value)
                    sh1.cell(row + 1, c, value=input(f"Enter the {m}: "))
                    column1 += 1
                wb.save("updated.xlsx")
                # break
            elif(choice == 2):
                movieNameToBeEdited = input("enter the movie name that you want to edit: ")
                for i in range(2, row + 1):
                    name = sh1.cell(i, column=1).value
                    if (movieNameToBeEdited == name):
                        for j in range(1, col + 1):
                                sh1.cell(i, j, value=input(f"Enter the {sh1.cell(1, j).value} new change: "))
                        # column1 += 1
                wb.save("MoviesInfo.xlsx")
            elif (choice == 3):
                movieToBeDeleted = input("Enter the movie name you want to Delete: ")
                for i in range(2, row + 1):
                    name = sh1.cell(i, column=1).value
                    if (name == movieToBeDeleted):
                        sh1.delete_rows(i, True)
                wb.save("MoviesInfo.xlsx")
                #break
            # elif (choice == 4):
            #     print("You have Successfully Logged out")
            #     sys.exit()


class UserDetails(AdminOfBookMyShow):

    def __init__(self, userName, userPassword):
        self.userCredentials= {
            userName : userPassword
        }
        print(f"*****Hello {self.userCredentials},you have Logged in Successfully*****")

    def userRegistration(self,username,password):
        if username in self.userCredentials.keys() and password == self.userCredentials[username]:
            print("login Successfull")
        else:
            print("login Failed")






    def ticketsBooking(self, choiceByUser):
        row, col, sh1, wb = workbook("MoviesInfo.xlsx", "Sheet1")
        for i in range(1, col + 1):
            print(sh1.cell(1, i).value, end=": ")
            print(sh1.cell(choiceByUser + 1, i).value)
        while(True):
            print("1.Book Tickets\n2. Cancel Ticket \n3. Give User Rating \n4.exit")
            userTicketChoice = int(input("Enter the choice from above: "))
            if (userTicketChoice == 1):
                print("***** Welcome User *****")
                for i in range(7, 9):
                    print(f"Timing : {sh1.cell(choiceByUser + 1, i).value}")
                timing = int(input("enter the timing you want to choose: "))
                print(f"selected timing {timing}")
                print(f"Remaining Seats is {sh1.cell(choiceByUser + 1, 13).value}")
                totalTicketsToBeBooked = int(input("Enter the number of tickets to be Booked:"))
                ticketsAvailable = sh1.cell(choiceByUser + 1, 13).value
                sh1.cell(choiceByUser + 1, 13, value=ticketsAvailable - totalTicketsToBeBooked)
                print(sh1.cell(choiceByUser + 1, 13).value)
                print("***** Thank You for Booking Tickets *****")
                wb.save("MoviesInfo.xlsx")
            elif (userTicketChoice == 2):
                ticketsTobeCancelled = int(input("enter the Number of tickets to be Cancelled"))
                ticketsAvailable = sh1.cell(choiceByUser + 1, 13).value
                sh1.cell(choiceByUser + 1, 13, value=ticketsAvailable + ticketsTobeCancelled)
                print(sh1.cell(choiceByUser + 1, 13).value)
                wb.save("MoviesInfo.xlsx")
            elif (userTicketChoice == 3):
                userRating = input("Enter your Review for above movie")
            elif(userTicketChoice==4):
                break


    def userActions(self):
        row, col, sh1, wb = workbook("MoviesInfo.xlsx", "Sheet1")
        m = 1
        for i in range(2, row + 1):
            print(m, end='.')
            print(sh1.cell(i, 1).value)
            m += 1
        print("4.Logout")

        choiceByUser = int(input("Enter the choice you want to choose"))
        if (choiceByUser == 1):
            print("***** Welcome User *****")
            self.ticketsBooking(1)
        elif(choiceByUser==2):
            print("***** Welcome User *****")
            self.ticketsBooking(2)
        elif(choiceByUser==3):
            print("***** Welcome User *****")
            self.ticketsBooking(3)






while (True):
    print("******Welcome to BookMyShow*******\n1.Login\n2.Register new account\n3.Exit")
    user_input = int(input("Enter your choice: "))
    if(user_input==3):
        break
    elif (user_input == 1):
        print("Welcome\n1.AdminLogin\n2.UserLogin\n3.Exit")
        loginType = int(input("enter your choice"))
        if (loginType == 1):
            name = input("enter you name: ")
            password = input("enter your password: ")
            admin = AdminOfBookMyShow(name, password)
            admin.adminOperations()
        elif (loginType == 2):
            userName = input("enter your UserName: ")
            UserPassword = input("enter the password: ")
            keshav = UserDetails(userName, UserPassword)
            keshav.userRegistration(userName, UserPassword)
            keshav.userActions()


        #break
    elif (user_input == 2):
        userName = input("Enter Your Name")
        # userEmail = input("Enter Your Email")
        # userPhoneNumber = int(input("Enter Your PhoneNumber"))
        # userAge = input("Enter Your Age")
        userPassword = input("Enter Your Password")
        keshav = UserDetails(userName, userPassword)
        user_Name=input("Enter your username")
        user_Password = input("Enter your Password")
        keshav.userRegistration(userName, user_Password)
        keshav.userActions()
        # break
    else:
        sys.exit()
