import sys
import openpyxl


# next=sh1.cell(2,8).value
# showLength=sh1.cell(2,3).value
# sum=next+showLength
# sh1.cell(2,12,value=sum)
# wb.save("forTimings.xlsx")
# no_of_shows=sh1.cell(2,7).value
# print(no_of_shows)
# print(col)
# for i in range(1,no_of_shows+1):
#     for column in range(13,col+1):
#         prev=sh1.cell(2,column-1).value
#         len=sh1.cell(2,3).value
#         sum=prev+len
#         sh1.cell(2, column, value=sum)
#     no_of_shows = sh1.cell(2, 7).value
# wb.save("forTimings.xlsx")

def workbook(filename, sheetname):
    wb = openpyxl.load_workbook(filename)
    sh1 = wb[sheetname]
    row = sh1.max_row
    col = sh1.max_column
    return row, col, sh1, wb


def userdict(self):
    row, col, sh1, wb=workbook("forTimings.xlsx","Sheet1")
    userdetailsDict = {}
    for i in range(1, row + 2):
        cell_value_class = sh1.cell(i, 1).value
        cell_value_id = sh1.cell(i, 5).value
        userdetailsDict[cell_value_class] = cell_value_id
    return userdetailsDict


class AdminOfBookMyShow():

    def __init__(self, adminName, adminPassword):
        """This is a constructor for admin"""
        self.adminName = adminName
        self.adminPassword = adminPassword

    def adminOperations(self):
        """This Method is used for All Admin Operations."""
        while (True):
            print("******WelcomeAdmin*******\n1.Add New Movie Info\n2.Edit Movie\n3.Delete Movie\n4.Logout")
            choice = int(input("Enter the choice you want to perform: "))
            row, col, sh1, wb = workbook("forTimings.xlsx", "Sheet1")
            if (choice == 4):
                break
            if (choice == 1):
                print("enter the details of the movie: ")
                row1 = 1
                column1 = 1
                for c in range(1, col -2):
                    m = (sh1.cell(row1, column1).value)
                    sh1.cell(row + 1, c, value=input(f"Enter the {m}: "))
                    column1 += 1

                next = sh1.cell(row, 8).value
                print(next)
                showLength = sh1.cell(row, 3).value
                print(showLength)
                sum = float(next + showLength)
                sh1.cell(row, 12, value=sum)
                # wb.save("forTimings.xlsx")
                no_of_shows = int(sh1.cell(row, 7).value)
                print(no_of_shows)
                print(col)
                print(row)
                for rowupdated in range(row,row+1):
                    for i in range(1, no_of_shows + 1):
                        for column in range(13, col + 1):
                            prev = float(sh1.cell(rowupdated, column - 1).value)
                            len = float(sh1.cell(rowupdated, 3).value)
                            sum = float(prev + len)
                            sh1.cell(rowupdated, column, value=sum)
                        no_of_shows = sh1.cell(2, 7).value
                wb.save("forTimings.xlsx")
                # break
            elif (choice == 2):
                movieNameToBeEdited = input("enter the movie name that you want to edit: ")
                for i in range(2, row + 1):
                    name = sh1.cell(i, column=1).value
                    if (movieNameToBeEdited == name):
                            for j in range(1, col + 1):
                                sh1.cell(i, j, value=input(f"Enter the {sh1.cell(1, j).value} new change: "))
                            # column1 += 1
                    else:
                        print(f"Their is no movie called{movieNameToBeEdited} .please check again")
                wb.save("forTimings.xlsx")
            elif (choice == 3):
                movieToBeDeleted = input("Enter the movie name you want to Delete: ")
                for i in range(2, row + 1):
                    name = sh1.cell(i, column=1).value
                    if (name == movieToBeDeleted):
                        sh1.delete_rows(i, True)
                wb.save("forTimings.xlsx")
                # break
            # elif (choice == 4):
            #     print("You have Successfully Logged out")
            #     sys.exit()


class UserDetails(AdminOfBookMyShow):

    def validatinguserRegistration(self, username, password):
        """This Method is used for validating the user credentials"""
        userdetailsdict=userdict(self)
        if username in userdetailsdict.keys() and password == userdetailsdict[username]:
            print("login Successfull")
        else:
            print("login Failed!!! Please Enter valid login Credentials and relogin again")
            sys.exit()

    def ticketsBooking(self, choiceByUser):
        """This Method is used for Booking Transactions"""
        row, col, sh1, wb = workbook("forTimings.xlsx", "Sheet1")
        for i in range(1, col + 1):
            print(sh1.cell(1, i).value, end=": ")
            print(sh1.cell(choiceByUser + 1, i).value)
        while (True):
            print("1.Book Tickets\n2. Cancel Ticket \n3. Give User Rating \n4.exit")
            userTicketChoice = int(input("Enter the choice from above: "))
            if (userTicketChoice == 1):
                print("***** Welcome User *****")
                timingnext = sh1.cell(2, 12).value
                print(f"1.{timingprev}-{timingnext}")
                i = 2
                for ticketsbook in range(13, col + 1):
                    timingprev = sh1.cell(2, ticketsbook - 1).value
                    timingnext = sh1.cell(2, ticketsbook).value
                    print(f"{i}.{timingprev}-{timingnext}")
                    i = i + 1
                # for i in range(7, 9):
                #     print(f"Timing : {sh1.cell(choiceByUser + 1, i).value}")
                timing = int(input("enter the timing you want to choose: "))
                print(f"selected timing {timing}")
                print(f"Remaining Seats is {sh1.cell(choiceByUser + 1, 13).value}")
                totalTicketsToBeBooked = int(input("Enter the number of tickets to be Booked:"))
                ticketsAvailable = sh1.cell(choiceByUser + 1, 13).value
                sh1.cell(choiceByUser + 1, 13, value=ticketsAvailable - totalTicketsToBeBooked)
                print(sh1.cell(choiceByUser + 1, 13).value)
                print("***** Thank You for Booking Tickets *****")
                wb.save("forTimings.xlsx")
            elif (userTicketChoice == 2):
                ticketsTobeCancelled = int(input("enter the Number of tickets to be Cancelled"))
                ticketsAvailable = sh1.cell(choiceByUser + 1, 13).value
                sh1.cell(choiceByUser + 1, 13, value=ticketsAvailable + ticketsTobeCancelled)
                print(sh1.cell(choiceByUser + 1, 13).value)
                wb.save("forTimings.xlsx")
            elif (userTicketChoice == 3):
                row, col, sh1, wb = workbook("userreviews.xlsx", "Sheet1")
                userRating = input("Enter your Review for above movie")
                row, col, sh1, wb = workbook("MoviesInfo.xlsx", "Sheet1")
                sh1.cell(2,1,value=sh1.cell(choiceByUser+1,1).value)
                sh1.cell(2,2,value=userRating)
                wb.save("forTimings.xlsx")
            elif (userTicketChoice == 4):
                break

    def moviesAvailable(self):
        """This method gives the available movies from excel sheet"""
        row, col, sh1, wb = workbook("forTimings.xlsx", "Sheet1")
        m = 1
        for i in range(2, row + 1):
            print(m, end='.')
            print(sh1.cell(i, 1).value)
            m += 1
        print("4.Logout")

        choiceByUser = int(input("Enter the choice you want to choose"))
        if (choiceByUser == 1):
            print("***** Welcome User *****")
            self.ticketsBooking(1)
        elif (choiceByUser == 2):
            print("***** Welcome User *****")
            self.ticketsBooking(2)
        elif (choiceByUser == 3):
            print("***** Welcome User *****")
            self.ticketsBooking(3)


while (True):
    print("******Welcome to BookMyShow*******\n1.Login\n2.Register new account\n3.Exit")
    user_input = int(input("Enter your choice: "))
    if (user_input == 3):
        break
    elif (user_input == 1):
        print("Welcome\n1.AdminLogin\n2.UserLogin\n3.Exit")
        loginType = int(input("enter your Choice: "))
        if (loginType == 1):
            name = input("Enter you username: ")
            password = input("Enter your password: ")
            admin = AdminOfBookMyShow(name, password)
            admin.adminOperations()
        elif (loginType == 2):
            userName = input("enter your UserName: ")
            UserPassword = input("enter the password: ")
            keshav = UserDetails(userName, UserPassword)
            keshav.validatinguserRegistration(userName, UserPassword)
            keshav.moviesAvailable()

        # break
    elif (user_input == 2):
        row, col, sh1, wb = workbook("RegisteredUserDetails.xlsx", "Sheet1")
        for i in range(1, col + 1):
            sh1.cell(row + 1, i, value=input(f"Enter your {sh1.cell(1, i).value}"))
        wb.save("RegisteredUserDetails.xlsx")
    else:
        sys.exit()
